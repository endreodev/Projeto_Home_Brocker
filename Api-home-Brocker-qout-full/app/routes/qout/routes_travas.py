from io import BytesIO
import requests
from flask import jsonify, request, abort, make_response, send_file
from flask_jwt_extended import get_jwt, jwt_required, get_jwt_identity
import openpyxl
from openpyxl.utils import get_column_letter
from app.database import db
from datetime import datetime, timedelta
from sqlalchemy import desc, extract, func
from app.functions.validacao_fin_semana import is_weekend
from app.models.models import Acessos, Empresa, Firebase, Integracao, Trava, Parceiro, User
import firebase_admin
from firebase_admin import credentials, messaging 
import pytz 
import os
import locale
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph,Spacer

from app.routes.rotas.integracao.integracao_sankhya import enviar_operacao

# Configura o locale para português do Brasil
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

def format_number(number):
    # Formata o número com separador de milhar e duas casas decimais
    return locale.format_string('%.2f', number, grouping=True)


cred_path = os.path.join(os.path.dirname(__file__), '../../../homebrokerdtvm-7824e290ff8d.json')
cred = credentials.Certificate(os.path.abspath(cred_path))


# cred = credentials.Certificate("../homebrokerdtvm-7824e290ff8d.json")
firebase_admin.initialize_app(cred)

def init_routes_trava(app):
    
    @app.route('/trava', methods=['POST'])
    @jwt_required()
    def create_trava():

        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)
        empresa_id = additional_claims.get('empresa_id', None)

        data = request.get_json()
        if not data:
            return jsonify({'message': 'Nenhum dado fornecido.'}), 400

        required_fields = ['empresa_id','parceiro_id','quantidade', 'preco_unitario', 'parceiro_id']
        if not all(data.get(field) for field in required_fields):
            return jsonify({'message': 'Faltando campos obrigatórios ou campos vazios'}), 400
        
        parc = Parceiro.query.get(data['parceiro_id'])
        if parc.lmt_trava < data['quantidade']:
            return  jsonify({'message': 'Quantidade da Ordem maior que o limite!'}), 400
        
        mes_atual = datetime.now().month
        ano_atual = datetime.now().year
        
        vendas_mes = db.session.query(func.sum(Trava.quantidade)).filter(
            Trava.parceiro_id == data['parceiro_id'],
            extract('month', Trava.created_at) == mes_atual,
            extract('year', Trava.created_at) == ano_atual
        ).scalar()

        if not vendas_mes:
                vendas_mes = 0
                
        #valida limite mensal
        if parc.lmt_mes < (vendas_mes+ int(data['quantidade']) ):
            disponivel = parc.lmt_mes - vendas_mes
            return jsonify({'message': 'O Limite mensal foi atingido! Disponível: {} Gramas'.format(disponivel)}), 400

        emp = Empresa.query.get(empresa_id)
        if not check_time_within(emp.hrinicio , emp.hrfinal):
            return jsonify({'message': 'Operação fora de horario permitido!'}), 400
        

        if is_weekend():
            return jsonify({'message': 'Operação proibida aos fins de semana!'}), 400
        
        new_trava = Trava(
            empresa_id=data['empresa_id'],
            parceiro_id=data['parceiro_id'],
            usuario_id=user_id,
            quantidade=data['quantidade'],
            preco_unitario=data['preco_unitario'],
            preco_total=data['preco_total'],
            cotacao=data['cotacao'], 
            dollar=data['dollar'], 
            desagio=parc.plano  
        )

        db.session.add(new_trava)
        try:
            db.session.commit()

            try:
                user = User.query.get(user_id)
                parceiro = Parceiro.query.get(data['parceiro_id'])
                data_hora_atual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                title = f' 🛒 Venda Realizada \n {data_hora_atual}  \n\nUsuario: {user.nome} \nParceiro: {parceiro.nome} \n'
                message = f'Quantidade:  {format_number(data["quantidade"])}, \nCotação: {data["preco_unitario"]}, \nTotal: {format_number(data["preco_total"])}'
                
                # print(message)
                disparo_de_notificacao( title , message , data['empresa_id'] )


            except Exception as ea:
                print("erro na notificação")

            return jsonify({'message': 'trava criada com sucesso!'}), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erro ao salvar no banco de dados', 'details': str(e)}), 500
    

    # Consultar todas as Trava 
    @app.route('/trava', methods=['GET'])
    @jwt_required()
    def get_travas_parceiro_todos():
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        data_inicial = request.args.get('data_inicial', None)
        data_final = request.args.get('data_final', None)
        
        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)

        # Consulta para buscar os parceiros aos quais o usuário tem acesso
        acessos = Acessos.query.filter_by(usuario_id=user_id, ativo=True).all()
        if not acessos:
            return jsonify({'message': 'Nenhum acesso encontrado para este usuário.'}), 404

        # Cria uma lista dos IDs dos parceiros
        parceiro_ids = [acesso.parceiro_id for acesso in acessos]

        # Consulta inicial
        query = db.session.query(
            Trava, 
            User.nome.label('nome_usuario'), 
            Parceiro.nome.label('nome_parceiro')
        ).join(User, Trava.usuario_id == User.id
        ).join(Parceiro, Trava.parceiro_id == Parceiro.id
        ).filter(Trava.parceiro_id.in_(parceiro_ids))

         # Filtros de data, se fornecidos
        if data_inicial:
            data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d')
            query = query.filter(Trava.created_at >= data_inicial)
            
        if data_final:
            data_final = datetime.strptime(data_final, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            query = query.filter(Trava.created_at <= data_final)

        # Ordenação e paginação
        travas = query.order_by(desc(Trava.id)).paginate(page=page, per_page=per_page, error_out=False)

        travas_list = [
            {   
                'id': trava.Trava.id,
                'empresa_id': trava.Trava.empresa_id,
                'parceiro_id': trava.Trava.parceiro_id,
                'parceiro_nome': trava.nome_parceiro,
                'usuario_id': trava.Trava.usuario_id,
                'usuario_nome': trava.nome_usuario,
                'produto_id': trava.Trava.produto_id,
                'quantidade': trava.Trava.quantidade,
                'preco_unitario': trava.Trava.preco_unitario,
                'preco_total': trava.Trava.preco_total,
                'status': trava.Trava.status,
                'integrado': trava.Trava.integrado,
                'dollar': trava.Trava.dollar,
                'created_at': format_datetime_br(trava.Trava.created_at),
                'updated_at': format_datetime_br(trava.Trava.updated_at),
                'integracao': serialize_integracao(
                    Integracao.query.filter_by(id_trava=trava.Trava.id).first()
                )
            } for trava in travas.items
        ]
        
        return jsonify({
            'travas': travas_list,
            'total': travas.total,
            'pages': travas.pages,
            'current_page': page
        })



    @app.route('/trava/parceiro/<int:id>', methods=['GET'])
    @jwt_required()
    def get_travas_parceiro(id):
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        data_inicial = request.args.get('data_inicial', None)
        data_final = request.args.get('data_final', None)
        
        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)

        # Consulta para buscar os parceiros aos quais o usuário tem acesso
        acessos = Acessos.query.filter_by(usuario_id=user_id, ativo=True).all()
        if not acessos:
            return jsonify({'message': 'Nenhum acesso encontrado para este usuário.'}), 404

        # Cria uma lista dos IDs dos parceiros
        parceiro_ids = [acesso.parceiro_id for acesso in acessos]
        existe_id = any(d == id for d in parceiro_ids)
        if not existe_id:
            return jsonify({'message': 'Sem acesso ao Parceiro'}), 400

        # Consulta inicial
        query = db.session.query(
            Trava, 
            User.nome.label('nome_usuario'), 
            Parceiro.nome.label('nome_parceiro')
        ).join(User, Trava.usuario_id == User.id
        ).join(Parceiro, Trava.parceiro_id == Parceiro.id
        ).filter(Trava.parceiro_id == id)

        # Filtros de data, se fornecidos
        if data_inicial:
            data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d')
            query = query.filter(Trava.created_at >= data_inicial)
            
        if data_final:
            data_final = datetime.strptime(data_final, '%Y-%m-%d')
            query = query.filter(Trava.created_at <= data_final)

        # Ordenação e paginação
        travas = query.order_by(desc(Trava.id)).paginate(page=page, per_page=per_page, error_out=False)

        travas_list = [
            {   
                'id': trava.Trava.id,
                'empresa_id': trava.Trava.empresa_id,
                'parceiro_id': trava.Trava.parceiro_id,
                'parceiro_nome': trava.nome_parceiro,
                'usuario_id': trava.Trava.usuario_id,
                'usuario_nome': trava.nome_usuario,
                'produto_id': trava.Trava.produto_id,
                'quantidade': trava.Trava.quantidade,
                'preco_unitario': trava.Trava.preco_unitario,
                'preco_total': trava.Trava.preco_total,
                'status': trava.Trava.status,
                'created_at': format_datetime_br(trava.Trava.created_at),
                'updated_at': format_datetime_br(trava.Trava.updated_at)
            } for trava in travas.items
        ]
        
        return jsonify({
            'travas': travas_list,
            'total': travas.total,
            'pages': travas.pages,
            'current_page': page
        })

    # consultar trava Trava
    @app.route('/trava/<int:id>', methods=['GET'])
    @jwt_required()
    def get_trava_id(id):

        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)

        # Consulta para buscar os parceiros aos quais o usuário tem acesso
        acessos = Acessos.query.filter_by(usuario_id=user_id, ativo=True).all()
        if not acessos:
            return jsonify({'message': 'Nenhum acesso encontrado para este usuário.'}), 404

        # Cria uma lista dos IDs dos parceiros
        parceiro_ids = [acesso.parceiro_id for acesso in acessos]
        existe_id = any(d == id for d in parceiro_ids)
        if not existe_id:
            return jsonify({'message': 'Sem acesso ao Parceiro'}), 400
        
        trava = Trava.query.get(id)
        if not trava:
            return jsonify({'message': 'trava não encontrada.'}), 404
        
        return jsonify({
                'id': trava.id,
                'empresa_id': trava.empresa_id,
                'parceiro_id': trava.parceiro_id,
                'usuario_id': trava.usuario_id,
                'produto_id': trava.produto_id,
                'quantidade': trava.quantidade,
                'preco_unitario': trava.preco_unitario,
                'preco_total': trava.preco_total,
                'status': trava.status,
                'created_at': format_datetime_br(trava.created_at),
                'updated_at': format_datetime_br(trava.updated_at)
        })

    # auterar unica Trava
    @app.route('/trava/<int:id>', methods=['PUT'])
    @jwt_required()
    def update_trava(id):

        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)

        # Consulta para buscar os parceiros aos quais o usuário tem acesso
        acessos = Acessos.query.filter_by(usuario_id=user_id, ativo=True).all()
        if not acessos:
            return jsonify({'message': 'Nenhum acesso encontrado para este usuário.'}), 404

        # Cria uma lista dos IDs dos parceiros
        parceiro_ids = [acesso.parceiro_id for acesso in acessos]
        existe_id = any(d == id for d in parceiro_ids)
        if not existe_id:
            return jsonify({'message': 'Sem acesso ao Parceiro'}), 400
    
        trava = Trava.query.get(id)
        if not trava:
            return jsonify({'message': 'trava não encontrada.'}), 404 

        data = request.get_json()
        trava.produto_id = data.get('produto_id', trava.produto_id)
        trava.quantidade = data.get('quantidade', trava.quantidade)
        trava.preco_unitario = data.get('preco_unitario', trava.preco_unitario)
        trava.parceiro_id = data.get('parceiro_id', trava.parceiro_id)

        try:
            db.session.commit()
            return jsonify({'message': 'trava atualizada com sucesso!'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erro ao atualizar a trava', 'details': str(e)}), 500
    
    
    # exclir trava 
    @app.route('/trava/<int:id>', methods=['DELETE'])
    @jwt_required()
    def delete_trava(id):
        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)

        # Consulta para buscar os parceiros aos quais o usuário tem acesso
        acessos = Acessos.query.filter_by(usuario_id=user_id, ativo=True).all()
        if not acessos:
            return jsonify({'message': 'Nenhum acesso encontrado para este usuário.'}), 404

        # Cria uma lista dos IDs dos parceiros
        parceiro_ids = [acesso.parceiro_id for acesso in acessos]
        existe_id = any(d == id for d in parceiro_ids)
        if not existe_id:
            return jsonify({'message': 'Sem acesso ao Parceiro'}), 400
        
        trava = Trava.query.get(id)
        if not trava:
            return jsonify({'message': 'trava não encontrada.'}), 404 

        try:
            db.session.delete(trava)
            db.session.commit()
            return jsonify({'message': 'trava deletada com sucesso!'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erro ao deletar a trava', 'details': str(e)}), 500


    # consultar trava Trava
    @app.route('/trava/encerrar/<int:id>', methods=['GET'])
    @jwt_required()
    def get_trava_encerrar_id(id):

        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)

        # Consulta para buscar os parceiros aos quais o usuário tem acesso
        acessos = Acessos.query.filter_by(usuario_id=user_id, ativo=True).all()
        if not acessos:
            return jsonify({'message': 'Nenhum acesso encontrado para este usuário.'}), 404

        # Cria uma lista dos IDs dos parceiros
        # parceiro_ids = [acesso.parceiro_id for acesso in acessos]
        # existe_id = any(d == id for d in parceiro_ids)
        # if not existe_id:
        #     return jsonify({'message': 'Sem acesso ao Parceiro'}), 400
        
        trava = Trava.query.get(id)
        if not trava:
            return jsonify({'message': 'Ordem não encontrada.'}), 404 

        trava.status = 'F'

        try:
            db.session.commit()
            return jsonify({'message': 'Ordem ENCERRADA com sucesso!'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erro ao atualizar a Ordem', 'details': str(e)}), 500

    # consultar trava Trava
    @app.route('/trava/cancelar/<int:id>', methods=['GET'])
    @jwt_required()
    def get_trava_cancelar_id(id):

        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)

        # Consulta para buscar os parceiros aos quais o usuário tem acesso
        acessos = Acessos.query.filter_by(usuario_id=user_id, ativo=True).all()
        if not acessos:
            return jsonify({'message': 'Nenhum acesso encontrado para este usuário.'}), 404

        # Cria uma lista dos IDs dos parceiros
        # parceiro_ids = [acesso.parceiro_id for acesso in acessos]
        # existe_id = any(d == id for d in parceiro_ids)
        # if not existe_id:
        #     return jsonify({'message': 'Sem acesso ao Parceiro'}), 400
        
        trava = Trava.query.get(id)
        if not trava:
            return jsonify({'message': 'Ordem não encontrada.'}), 404 

        trava.status = 'C'

        try:
            db.session.commit()
            return jsonify({'message': 'Ordem CANCELADA com sucesso!'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Erro ao atualizar a Ordem', 'details': str(e)}), 500


    @app.route('/trava/relatorio', methods=['GET'])
    @jwt_required()
    def gerar_relatorio():
        # Obtenha os parâmetros de paginação e datas
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 1000, type=int)
        data_inicial = request.args.get('data_inicial', None)
        data_final = request.args.get('data_final', None)

        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)
        empresa_id = additional_claims.get('empresa_id', None)

        # Consulta para buscar os parceiros aos quais o usuário tem acesso
        acessos = Acessos.query.filter_by(usuario_id=user_id, ativo=True).all()
        if not acessos:
            return jsonify({'message': 'Nenhum acesso encontrado para este usuário.'}), 404

        parceiro_ids = [acesso.parceiro_id for acesso in acessos]

        # Consulta inicial
        query = db.session.query(
            Trava, 
            User.nome.label('nome_usuario'), 
            Parceiro.nome.label('nome_parceiro')
        ).join(User, Trava.usuario_id == User.id
        ).join(Parceiro, Trava.parceiro_id == Parceiro.id
        ).filter(Trava.parceiro_id.in_(parceiro_ids))

        # Filtros de data, se fornecidos
        if data_inicial:
            data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d')
            query = query.filter(Trava.created_at >= data_inicial)
        else:
            data_inicial = "Não informado"
            
        if data_final:
            data_final = datetime.strptime(data_final, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            query = query.filter(Trava.created_at <= data_final)
        else:
            data_final = "Não informado"

        # Ordenação e paginação
        travas = query.order_by(desc(Trava.id)).paginate(page=page, per_page=per_page, error_out=False)

        # Usar BytesIO como buffer para o PDF
        buffer = BytesIO()
        
        # Criar o PDF
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        nome_empresa = Empresa.query.get(empresa_id).nome
        
        # Adicionar o nome da empresa ao relatório
        styles = getSampleStyleSheet()
        empresa_nome = nome_empresa
        empresa_paragraph = Paragraph(f"<b>{empresa_nome}</b>", styles['Title'])
        elements.append(empresa_paragraph)

        # Adicionar os filtros ao relatório
        filtros_paragraph = Paragraph(
            f"Período: <b>{data_inicial}</b> a <b>{data_final}</b>",
            styles['Normal']
        )
        elements.append(filtros_paragraph)

        # Adicionar um espaçamento após os filtros
        elements.append(Spacer(1, 12))

        # Adicionar um título ao relatório
        title = Paragraph("Relatório de Travas", styles['Title'])
        elements.append(title)

        # Construa os dados para a tabela
        data = [["ID", "Parceiro", "Usuário", "Quantidade", "Preço Unitário", "Total", "Status", "Data"]]

        total_quantidade = 0
        total_preco_unitario = 0
        total_preco_total = 0

        for trava in travas.items:
            quantidade = trava.Trava.quantidade
            preco_unitario = trava.Trava.preco_unitario
            preco_total = trava.Trava.preco_total

            total_quantidade += quantidade
            total_preco_unitario += preco_unitario
            total_preco_total += preco_total

            data.append([
                trava.Trava.id,
                trava.nome_parceiro,
                trava.nome_usuario,
                quantidade,
                f"R$ {preco_unitario:,.2f}",
                f"R$ {preco_total:,.2f}",
                trava.Trava.status,
                format_datetime_br(trava.Trava.created_at)
            ])
        

        # Adicionar totalizadores ao final da tabela
        data.append([
            "Totais",
            "",  # Deixar as colunas que não serão somadas em branco
            "",
            f"{total_quantidade}",
            f"R$ {total_preco_unitario:,.2f}",
            f"R$ {total_preco_total:,.2f}",
            "",
            ""
        ])

        # Estilizar a tabela
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)

        # Construir o PDF no buffer
        doc.build(elements)

        # Enviar o conteúdo do PDF como resposta
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name='relatorio_travas.pdf', mimetype='application/pdf')
    

    @app.route('/trava/relatorio_excel', methods=['GET'])
    @jwt_required()
    def gerar_relatorio_excel():
        # Obtenha os parâmetros de paginação e datas
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        data_inicial = request.args.get('data_inicial', None)
        data_final = request.args.get('data_final', None)

        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)

        # Consulta para buscar os parceiros aos quais o usuário tem acesso
        acessos = Acessos.query.filter_by(usuario_id=user_id, ativo=True).all()
        if not acessos:
            return jsonify({'message': 'Nenhum acesso encontrado para este usuário.'}), 404

        parceiro_ids = [acesso.parceiro_id for acesso in acessos]

        # Consulta inicial
        query = db.session.query(
            Trava, 
            User.nome.label('nome_usuario'), 
            Parceiro.nome.label('nome_parceiro')
        ).join(User, Trava.usuario_id == User.id
        ).join(Parceiro, Trava.parceiro_id == Parceiro.id
        ).filter(Trava.parceiro_id.in_(parceiro_ids))

        # Filtros de data, se fornecidos
        if data_inicial:
            data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d')
            query = query.filter(Trava.created_at >= data_inicial)
        else:
            data_inicial = "Não informado"
            
        if data_final:
            data_final = datetime.strptime(data_final, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            query = query.filter(Trava.created_at <= data_final)
        else:
            data_final = "Não informado"

        # Ordenação e paginação
        travas = query.order_by(desc(Trava.id)).paginate(page=page, per_page=per_page, error_out=False)

        # Criar o arquivo Excel usando openpyxl
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Relatório de Travas'

        # Adicionar o nome da empresa e filtros no topo
        empresa_nome = "Nome da Empresa"
        sheet["A1"] = empresa_nome
        sheet["A2"] = f"Período: {data_inicial} a {data_final}"

        # Títulos da tabela
        headers = ["ID", "Parceiro", "Usuário", "Quantidade", "Preço Unitário", "Total", "Status", "Data"]
        sheet.append(headers)

        # Adicionar os dados à planilha
        total_quantidade = 0
        total_preco_unitario = 0
        total_preco_total = 0

        for trava in travas.items:
            quantidade = trava.Trava.quantidade
            preco_unitario = trava.Trava.preco_unitario
            preco_total = trava.Trava.preco_total

            total_quantidade += quantidade
            total_preco_unitario += preco_unitario
            total_preco_total += preco_total

            sheet.append([
                trava.Trava.id,
                trava.nome_parceiro,
                trava.nome_usuario,
                quantidade,
                preco_unitario,
                preco_total,
                trava.Trava.status,
                format_datetime_br(trava.Trava.created_at)
            ])

        # Adicionar totalizadores no final
        sheet.append(["Totais", "", "", total_quantidade, total_preco_unitario, total_preco_total, "", ""])

        # Ajustar a largura das colunas
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 15

        # Salvar o arquivo Excel no buffer
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Enviar o arquivo Excel como resposta
        return send_file(output, as_attachment=True, download_name='relatorio_travas.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # return send_file(buffer, as_attachment=True, download_name='relatorio_travas.pdf', mimetype='application/pdf')

    # Consultar cenda do parceiro por mes a ano atual 
    @app.route('/trava/mes-parceiro/<int:parceiro_id>', methods=['GET'])
    @jwt_required()
    def get_vendas_parceiro_mes(parceiro_id):
        mes_atual = datetime.now().month
        ano_atual = datetime.now().year

        additional_claims = get_jwt()
        user_id = additional_claims.get('user_id', None)

        # Consulta para buscar os parceiros aos quais o usuário tem acesso
        acessos = Acessos.query.filter_by(usuario_id=user_id, ativo=True).all()
        if not acessos:
            return jsonify({'message': 'Nenhum acesso encontrado para este usuário.'}), 404

        # Cria uma lista dos IDs dos parceiros
        parceiro_ids = [acesso.parceiro_id for acesso in acessos]
        existe_id = any(d == parceiro_id for d in parceiro_ids)
        if not existe_id:
            return jsonify({'message': 'Sem acesso ao Parceiro'}), 400 
        
        parc = Parceiro.query.get(parceiro_id)
        if not parc:
            return jsonify({'message': 'Parceiro não encontrado'}), 400
        
        vendas = db.session.query(func.sum(Trava.quantidade)).filter(
            Trava.parceiro_id == parceiro_id,
            extract('month', Trava.created_at) == mes_atual,
            extract('year', Trava.created_at) == ano_atual
        ).scalar()
        
        return jsonify({
            'parceiro_id': parceiro_id,
            'mes': mes_atual,
            'ano': ano_atual,
            'limite': parc.lmt_mes,
            'quantidade_vendas': vendas
        })

    # realiza integração
    @app.route('/trava/trava-integracao/<int:id>', methods=['GET']) 
    def get_trava_integra(id):

        dataIntegracao = Integracao.query.filter_by(id_trava=id).all() 
        # Verifica se a consulta retornou algum registro
        if dataIntegracao:
            return {"error": True, "message": "Registro já integrado!"}
        # Utilize esta função passando o ID da trava para testar a integração
        resultado = enviar_operacao(id)  # Suponha que 123 é o ID de uma trava existente

        return jsonify(resultado)
    
    # consulta trava 
    @app.route('/trava/trava-integracao-get/<int:id>', methods=['GET']) 
    def get_trava_integra_id_trava(id):

        dataIntegracao = Integracao.query.filter_by(id_trava=id).first()
        # Verifica se a consulta retornou algum registro
        if not dataIntegracao:
            return {"error": True, "message": "Registro não integrado!"}
        

        return jsonify(dataIntegracao.to_integracao())

def format_datetime_br(date):
    return date.strftime('%d/%m/%Y %H:%m') if date else ''


def disparo_de_notificacao(title, message, empresa_id):

    users = Firebase.query.filter_by(interno=True, empresa_id=empresa_id).all()
    for user in users:
        send_push_notification(user.token, title, message)

    message = title +"\n "+message
    enviar_mensagem_telegram(message, empresa_id)


def send_push_notification(token, title, body):
    try:
        message = messaging.Message(
            notification=messaging.Notification(title=title, body=body),
            token=token
        )
        response = messaging.send(message)  # Asegure-se de que sua biblioteca de mensagens é assíncrona
        print('Successfully sent message:', response)
    except Exception as e:
        print('Failed to send message:', e) #



def enviar_mensagem_telegram(mensagem, empresa_id):

    emp = Empresa.query.get(empresa_id)
    # print(empresa_id)
    if emp and emp.tokenbot: 
        url = f"https://api.telegram.org/bot{emp.tokenbot}/sendMessage"
    else: 
        print("Empresa não encontrada ou token do bot não configurado")
        return 

    # url = f"https://api.telegram.org/bot7396251711:AAEC9Iy4uy2bNiup3pNC3MP6_eSRkGEkX-k/sendMessage"
    # emp = Empresa.query.get(1)
    
    if not emp.telegran:
        print('Não configurado para telegran') #
        return 

    dados = {
        "chat_id": emp.telegran, #1534370927, #
        "text": mensagem
    }

    resposta = requests.post(url, data=dados)
    # print(resposta) 
    return


# def enviar_mensagem_telegram2(mensagem):
#     url = f"https://api.telegram.org/bot7396251711:AAEC9Iy4uy2bNiup3pNC3MP6_eSRkGEkX-k/sendMessage"

#     dados = {
#         "chat_id": 1534370927,
#         "text": mensagem
#     }

#     resposta = requests.post(url, data=dados)
#     print(resposta) 
#     return



def check_time_within(hrinicio , hrfinal):
    # Configuração de fuso horário, ajuste para o fuso horário necessário
    tz = pytz.timezone('America/Sao_Paulo')
    
    # Obtenção da hora atual no fuso horário correto
    now = datetime.now(tz)
    current_time = now.strftime('%H%M')

    # Convertendo os horários de string para int
    start_time = int(hrinicio)
    end_time = int(hrfinal)
    current_time = int(current_time)

    # Verificação se o horário atual está dentro do intervalo
    if start_time <= current_time <= end_time:
        return True
    else:
        return False
    

def serialize_integracao(integracao):
    if integracao is None:
        return None
    return {
        'id'         : integracao.id          ,
        'id_trava'   : integracao.id_trava    ,
        'codcontrole': integracao.codcontrole ,
        'sucesso'    : integracao.sucesso     ,
        'mensagem'   : integracao.mensagem    , 
        'created_at' : integracao.created_at  ,
        'updated_at' : integracao.updated_at  
    }


# def check_time_within(hrinicio, hrfinal):
#     # Se hrinicio ou hrfinal forem nulos, retorne True
#     if hrinicio is None or hrfinal is None:
#         return True

#     # Configuração de fuso horário, ajuste para o fuso horário necessário
#     tz = pytz.timezone('America/Sao_Paulo')
    
#     # Obtenção da hora atual no fuso horário correto
#     now = datetime.now(tz)
#     current_time = int(now.strftime('%H%M'))

#     # Convertendo os horários de string para int
#     start_time = int(hrinicio)
#     end_time = int(hrfinal)

#     # Verificação se o horário atual está dentro do intervalo
#     return start_time <= current_time <= end_time



